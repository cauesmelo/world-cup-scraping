__author__ = "Cauê Melo"
__credits__ = ["Cauê Melo", "Morganna Carmem Diniz"]
__license__ = "MIT"
__version__ = "0.0.1"
__maintainer__ = "Cauê Melo"
__email__ = "caue.melo@uniriotec.br"
__status__ = "Development"

# Desempenho do Brasil nas copas do mundo

# libs
from bs4 import BeautifulSoup
import requests
import re
from openpyxl import Workbook
from enum import IntEnum

# Change this to change script for another country :)
# US-FIFA standard must be used
# for reference: https://www.fifa.com/associations/
COUNTRY = "Brazil"


# DON'T CHANGE THIS
# it is used for insensitive comparison
COUNTRY = COUNTRY.lower()

# global variable for assign cup_id
cup_id = 0

#global pointer for rows in table
cup_info_row = 1
match_info_row = 1
goal_info_row = 1
rank_info_row = 1

# enum for table columns index
class CUP(IntEnum):
  CUP_ID = 1
  HOST = 2
  YEAR = 3
  FINAL_PHASE = 4
  WINNER = 5

class MATCH(IntEnum):
  GAME_ID = 1
  CUP_ID = 2
  VERSUS = 3
  PHASE = 4
  REFEREE = 5
  REFEREE_NAC = 6
  STADIUM = 7
  VENUE = 8
  TIME = 9
  RESULT = 10
  GOAL_PRO = 11
  GOAL_CON = 12

class GOAL(IntEnum):
  MATCH_ID = 1
  CUP_ID = 2
  PLAYER = 3
  TIME = 4
  PRO = 5

class RANK(IntEnum):
  RANK_ID = 1
  DATE = 2
  RANK = 3
  TEAM = 4
  TOTAL_POINTS = 5
  PREVIOUS_POINTS = 6
  RANK_DIFF = 7


# table & sheet creation
workbook = Workbook()
worksheet = workbook.active
worksheet.title = "cup_info"
workbook.create_sheet("match_info")
workbook.create_sheet("goal_info")
workbook.create_sheet("rank_info")
workbook.active = 1

worksheet.cell(cup_info_row, column = CUP.CUP_ID).value = "cup_id"
worksheet.cell(cup_info_row, column = CUP.HOST).value = "host"
worksheet.cell(cup_info_row, column = CUP.YEAR).value = "year"
worksheet.cell(cup_info_row, column = CUP.FINAL_PHASE).value = "final_phase"
worksheet.cell(cup_info_row, column = CUP.WINNER).value = "winner"

worksheet = workbook["match_info"]
worksheet.cell(match_info_row, column = MATCH.GAME_ID).value = "match_id"
worksheet.cell(match_info_row, column = MATCH.CUP_ID).value = "cup_id"
worksheet.cell(match_info_row, column = MATCH.VERSUS).value = "versus"
worksheet.cell(match_info_row, column = MATCH.PHASE).value = "phase"
worksheet.cell(match_info_row, column = MATCH.REFEREE).value = "referee"
worksheet.cell(match_info_row, column = MATCH.REFEREE_NAC).value = "referee_nac"
worksheet.cell(match_info_row, column = MATCH.STADIUM).value = "stadium"
worksheet.cell(match_info_row, column = MATCH.VENUE).value = "venue"
worksheet.cell(match_info_row, column = MATCH.TIME).value = "time"
worksheet.cell(match_info_row, column = MATCH.RESULT).value = "result"
worksheet.cell(match_info_row, column = MATCH.GOAL_PRO).value = "goal_pro"
worksheet.cell(match_info_row, column = MATCH.GOAL_CON).value = "goal_con"

worksheet = workbook["goal_info"]
worksheet.cell(goal_info_row, column = GOAL.MATCH_ID).value = "match_id"
worksheet.cell(goal_info_row, column = GOAL.CUP_ID).value = "cup_id"
worksheet.cell(goal_info_row, column = GOAL.PLAYER).value = "player"
worksheet.cell(goal_info_row, column = GOAL.TIME).value = "time"
worksheet.cell(goal_info_row, column = GOAL.PRO).value = "pro"

worksheet = workbook["rank_info"]
worksheet.cell(rank_info_row, column = RANK.RANK_ID).value = "rank_id"
worksheet.cell(rank_info_row, column = RANK.DATE).value = "date"
worksheet.cell(rank_info_row, column = RANK.RANK).value = "rank"
worksheet.cell(rank_info_row, column = RANK.TEAM).value = "team"
worksheet.cell(rank_info_row, column = RANK.TOTAL_POINTS).value = "total_points"
worksheet.cell(rank_info_row, column = RANK.PREVIOUS_POINTS).value = "previous_points"
worksheet.cell(rank_info_row, column = RANK.RANK_DIFF).value = "rank_diff"

# scrap fifa site for available editions
def fetch_editions():

  # request && parse
  main_page = BeautifulSoup(requests.get("https://www.fifa.com/worldcup/").content, 'html.parser')

  # scrap inside editions selector
  selector = str(main_page.find(id = "edition-selector"))
  cup_links = re.findall(r'value=\"(.*)\"', selector)
  cup_names = re.findall(r'\">\s*(.*)\n\s*</option>', selector)
  cup_editions = []

  # put all name and links for cup data inside an list of tuples
  # here i need to subtract 2 entries because it is future world cups
  for i in range(len(cup_names) - 2):
    cup_editions.append(tuple((cup_names.pop(), cup_links.pop())))
  return cup_editions


def scrap_goals(match_page, match_id, home):
  global goal_info_row
  worksheet = workbook["goal_info"]
  goals = match_page.find_all(class_ = "fi-mh__scorer")
  for goal in goals:
    goal_author = goal.find_all(class_ = "fi-p__nShorter")[0].contents[0]
    goal_time_raw = goal.find_all(class_ = "fi-mh__scorer__minute")[0].contents[0]
    if(goal_author != "@@shortname"):
      goal_time = re.findall(r'\s([0-9]*)\'', goal_time_raw)[0]
      if(goal.parent.parent['class'][0] == 'fi-mh__scorers__away' and home == False):
        goal_pro = True
      elif(goal.parent.parent['class'][0] == 'fi-mh__scorers__home' and home == True):
        goal_pro = True
      else:
        goal_pro = False
      
      #writing routines
      goal_info_row += 1
      worksheet.cell(goal_info_row, column = GOAL.CUP_ID).value = cup_id
      worksheet.cell(goal_info_row, column = GOAL.MATCH_ID).value = match_id
      worksheet.cell(goal_info_row, column = GOAL.PLAYER).value = goal_author
      worksheet.cell(goal_info_row, column = GOAL.TIME).value = goal_time
      worksheet.cell(goal_info_row, column = GOAL.PRO).value = goal_pro


def scrap_matches(link):
  matches_page = BeautifulSoup(requests.get("https://www.fifa.com" + link + "matches").content, 'html.parser')  
  matches = matches_page.find_all(class_ = "fi-mu result fi-mu-national result")

  for match in matches:
    team1 = match.find_all(class_ = "fi-t__nText")[0].contents[0]
    team2 = match.find_all(class_ = "fi-t__nText")[1].contents[0]
    link_match = match.parent['href']

    if(team1.lower() == COUNTRY or team2.lower() == COUNTRY):
      scrap_match(link_match)

def scrap_match(link):
  global match_info_row
  worksheet = workbook["match_info"]
  match_info_row += 1

  match_page = BeautifulSoup(requests.get("https://www.fifa.com" + link).content, 'html.parser')
  
  # scrap oponent
  teams = match_page.find_all(class_ = "fi-t__nText")

  if(teams[0].contents[0].lower() != COUNTRY):
    home = False
    versus = teams[0].contents[0]
  else:
    home = True
    versus = teams[1].contents[0]

  match_id = re.findall(r'match/([0-9]*)/', link)[0]
  stadium = match_page.find_all(class_ = "fi__info__stadium")[0].contents[0][:-1]
  venue = match_page.find_all(class_ = "fi__info__venue")[0].contents[0]
  preprocessed_time = match_page.find_all(class_ = "fi-mu__info__datetime")[0].contents[0]
  time = re.findall(r'^\s*(.*)\s*', preprocessed_time)[0]
  phase = match_page.find_all(class_ = "fi__info__group fi-ltr--force")[0].contents[0][:-1]

  # scrap score
  score = match_page.find_all(class_ = "fi-s__scoreText")[0].contents[0]
  home_score = re.findall(r'([0-9]*)\-', score)[0]
  away_score = re.findall(r'-([0-9]*)\b', score)[0]

  if(home):
    goal_pro = home_score
    goal_con = away_score
  else:
    goal_pro = away_score
    goal_con = home_score
  if(goal_pro > goal_con):
    result = "win"
  elif(goal_pro == goal_con):
    result = "draw"
  elif(goal_pro < goal_con):
    result = "lose"

  # writing routines
  worksheet.cell(match_info_row, column = MATCH.GAME_ID).value = match_id
  worksheet.cell(match_info_row, column = MATCH.CUP_ID).value = cup_id
  worksheet.cell(match_info_row, column = MATCH.VERSUS).value = versus
  worksheet.cell(match_info_row, column = MATCH.STADIUM).value = stadium
  worksheet.cell(match_info_row, column = MATCH.VENUE).value = venue
  worksheet.cell(match_info_row, column = MATCH.TIME).value = time
  worksheet.cell(match_info_row, column = MATCH.GOAL_PRO).value = goal_pro
  worksheet.cell(match_info_row, column = MATCH.GOAL_CON).value = goal_con
  worksheet.cell(match_info_row, column = MATCH.RESULT).value = result
  worksheet.cell(match_info_row, column = MATCH.PHASE).value = phase
  
  scrap_goals(match_page, match_id, home)
  

def scrap_cup(cup):
  global cup_info_row
  global cup_id
  cup_id += 1
  worksheet = workbook["cup_info"]
  cup_info_row += 1
  host = cup[0]
  year = re.findall(r'\s(.[0-9]*)\b', cup[0])[0]

  #writing routines
  worksheet.cell(cup_info_row, column = CUP.CUP_ID).value = cup_id
  worksheet.cell(cup_info_row, column = CUP.HOST).value = host[:-5]
  worksheet.cell(cup_info_row, column = CUP.YEAR).value = year

  #scrap matches for cup
  scrap_matches(cup[1])

def fetch_ranking_editions():
  ranking_page = BeautifulSoup(requests.get("https://www.fifa.com/fifa-world-ranking/ranking-table/men/").content, 'html.parser')
  ranking_updates_raw = ranking_page.find_all(class_="fi-ranking-schedule__nav__item")
  ranking_editions = []
  for ranking_update in ranking_updates_raw:
    ranking_link = ranking_update.find('a')
    print(ranking_link['href'])
    ranking_id = re.findall(r'/rank/id([0-9]*)/', ranking_link['href'])[0]
    ranking_editions.append((ranking_id, ranking_link.text, ranking_link['href']))

  return ranking_editions

def scrap_ranking(ranking_edition):
  print("xD")




# function calls
# cup_editions = fetch_editions()

# for cup in cup_editions:
#   scrap_cup(cup)

ranking_editions = fetch_ranking_editions()

for ranking in ranking_editions:
  scrap_ranking(ranking)


# save data
workbook.save("data.xlsx")
